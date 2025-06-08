import time
import json
import asyncio
import logging
import schedule
import openpyxl
import arabic_reshaper
from threading import Thread
from datetime import datetime
from reportlab.lib import colors
from ..utils.paths import FONTS_DIR
from reportlab.lib.units import inch
from bidi.algorithm import get_display
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import letter, landscape
from .utils import EmailService, DatabaseOps, create_folder
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    PageBreak,
)

try:
    from .paths import CHAT_HIST_DIR
    from .inference import inference_pipeline
except ImportError:
    from paths import CHAT_HIST_DIR
    from inference import inference_pipeline

logger = logging.getLogger(__name__)


class BaseMonitorService:
    """
    Base class for monitoring services using the observer pattern.
    Manages email subscriptions and periodic task scheduling.
    """

    def __init__(
        self,
        email_service: EmailService,
        start_service: bool = True,
        every_hours: int = 24,
    ):
        """
        Initialize the BaseMonitorService with email service and scheduling options.

        Args:
            email_service (EmailService): The email service instance to use for notifications.
            start_service (bool): Whether to start the monitoring service immediately. Defaults to True.
            every_hours (int): Interval in hours for scheduled tasks. Defaults to 24.
        """
        self.db = DatabaseOps()
        self.email_service = email_service
        self.start_service = start_service
        self.every_hours = every_hours
        self._running = False
        self._thread = None

        logger.info(f"{self.__class__.__name__} initialized with _running = False")

        self.email_service.subscribe(
            self._on_email_change
        )  # subscribe to the email service

        if self.start_service and self.email_service.receiver_email is not None:
            self._start_monitoring()

    def _on_email_change(self, old_email, new_email):
        """
        Callback method triggered when the email address changes.
        """
        logger.info(
            f"{self.__class__.__name__} Email changed from '{old_email}' to '{new_email}', running = {self._running}"
        )
        if not old_email and new_email:
            self._start_monitoring()
        elif not new_email and old_email:
            # waiting for the background thread to join is blocking so we need it to run async
            asyncio.create_task(self._stop_monitoring())

    def _start_monitoring(self):
        """
        Start a separate thread to monitor responses periodically.
        Uses a scheduler to execute tasks at specified intervals.
        """
        if self.start_service and not self._running:
            schedule.every(self.every_hours).hours.do(self._retrieve_and_email)
            self._thread = Thread(target=self._run_scheduler, daemon=False)
            self._running = True
            self._thread.start()

            logger.info(
                f"{self.__class__.__name__} thread started. Will run every {self.every_hours} hours."
            )

    def _stop_monitoring(self):
        """Stop the monitoring thread and clean up resources."""
        if self._running:
            logger.info(
                f"{self.__class__.__name__}: Stopping monitoring, setting _running to False"
            )
            schedule.clear()
            self._running = False
            if self._thread is not None:
                self._thread.join(timeout=5)  # Set a timeout to avoid hanging
                self._thread = None
            logger.info(f"{self.__class__.__name__}: thread stopped and cleaned up.")

    def _run_scheduler(self):
        logger.info(
            f"{self.__class__.__name__}: Scheduler loop starting with _running = True"
        )
        while self._running:
            # logger.info(f"{self.__class__.__name__}: Scheduler checking for pending jobs...")
            schedule.run_pending()
            time.sleep(1)

        logger.info(
            f"{self.__class__.__name__}: Scheduler loop exited because _running = False"
        )

    def _retrieve_and_email(self):
        """
        Abstract method to retrieve data and send email notifications.
        Must be implemented by subclasses.

        Raises:
            NotImplementedError: If not implemented by a subclass.
        """
        raise NotImplementedError("Subclasses must implement this method")


class UncertainResponseMonitor(BaseMonitorService):
    """
    Monitors and reports uncertain responses from LLMs using a classifier.
    Inherits from BaseMonitorService.
    """

    def _retrieve_and_email(self):
        unknowns = []
        q_a = self.db.get_monitored_resp()
        logger.info(f"Fetched {len(q_a)} responses from the last 24 hours.")

        for question, answer in q_a:
            pred = inference_pipeline(answer[:60])
            if pred[0] == 1:  # the LLM didn't know the answer
                unknowns.append((question, answer))

        if unknowns:
            logger.info(f"Found {len(unknowns)} uncertain responses. Sending email...")
            subject = f"RAG System: {len(unknowns)} Uncertain Responses Detected"
            self.email_service.send_email(subject, unknowns)
        else:
            logger.info("No uncertain responses found.")


class ChatHistoryMonitor(BaseMonitorService):
    """
    Monitors and emails chat history periodically.
    Inherits from BaseMonitorService.
    """

    def _retrieve_and_email(self):
        """
        Retrieve chat history, save it to files in multiple formats, and send via email.
        """
        now = datetime.now().strftime("%Y-%m-%d_%H-%M")
        receiver_email = self.email_service.receiver_email
        base_filename = f"{now}_{receiver_email}"

        history = self.db.get_chat_history(full_history=True, last_n_hours=24)

        if not history:
            logger.info(f"{self.__class__.__name__}: No chat history.")
            return

        output_dir = create_folder(CHAT_HIST_DIR / receiver_email)
        created_files = self._save_history_to_files(history, base_filename, output_dir)

        if not created_files:
            logger.warning(
                f"{self.__class__.__name__}: Failed to create any history files."
            )
            return

        if hasattr(self.email_service, "send_email_with_attachments"):
            file_paths = [str(f) for f in created_files]
            subject = "Chat History for Your Chatbot"
            message_body = (
                f"Please find attached the chat history from the past 24 hours.\n\n"
                f"Attached files:\n"
                f"- JSON: Raw conversation data\n"
                f"- XLSX: Spreadsheet format for analysis\n"
                f"- PDF: Formatted document for easy reading\n\n"
                f"Best Regards,\n"
                f"Chatbot Monitoring System"
            )

            self.email_service.send_email_with_attachments(
                subject=subject, message_body=message_body, file_paths=file_paths
            )
        else:
            logger.info(
                f"{self.__class__.__name__}: Found previous conversations. Sending email with JSON only..."
            )
            subject = "Chat History for Your Chatbot"

            json_file = next(
                (f for f in created_files if f.suffix.lower() == ".json"), None
            )
            if json_file:
                self.email_service.send_email(
                    subject=subject, json_data=history, filename=json_file.name
                )

    def _save_history_to_files(self, history, base_filename, output_dir):
        """
        Save chat history to multiple file formats.

        Args:
            history (list[dict]): The chat history data to save
            base_filename (str): Base filename without extension
            output_dir (Path): Directory to save files in

        Returns:
            list[Path]: List of created file paths
        """
        created_files = []

        json_file = output_dir / f"{base_filename}.json"
        if self._save_json(history, json_file):
            created_files.append(json_file)

            xlsx_file = output_dir / f"{base_filename}.xlsx"
            if self._create_xlsx(history, xlsx_file):
                created_files.append(xlsx_file)

            pdf_file = output_dir / f"{base_filename}.pdf"
            if self._create_pdf(history, pdf_file):
                created_files.append(pdf_file)

        return created_files

    def _save_json(self, data, output_path):
        """
        Save data to a JSON file.

        Args:
            data (list[dict]): The data to save
            output_path (Path): Path to the output file

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            with open(output_path, "w", encoding="utf-8") as outfile:
                json.dump(data, outfile, indent=4, ensure_ascii=False)
            logger.info(f"JSON file saved successfully: {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error saving JSON file: {e}")
            return False

    def _create_xlsx(self, data, output_path):
        """
        Create an Excel file from chat history data.

        Args:
            data (list[dict]): Chat history data
            output_path (Path): Path to the output XLSX file

        Returns:
            bool: True if successful, False otherwise
        """
        logger.info(f"Creating XLSX file: {output_path}")
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Conversations"

            current_row = 1

            for entry in data:
                user_id = entry.get("user_id", "")

                llm_value = ""
                embedder_value = ""
                if entry.get("interactions"):
                    first_interaction = entry["interactions"][0]
                    llm_value = first_interaction.get("llm", "")
                    embedder_value = first_interaction.get("embedder", "")

                ws.cell(row=current_row, column=1, value="user_id:")
                ws.cell(row=current_row, column=2, value=user_id)
                current_row += 1

                ws.cell(row=current_row, column=1, value="llm:")
                ws.cell(row=current_row, column=2, value=llm_value)
                current_row += 1

                ws.cell(row=current_row, column=1, value="embedder:")
                ws.cell(row=current_row, column=2, value=embedder_value)
                current_row += 1

                ws.cell(row=current_row, column=1, value="interactions:")
                current_row += 1

                for interaction in entry.get("interactions", []):
                    user_text = interaction.get("user", "")
                    assistant_text = interaction.get("assistant", "")
                    timestamp_text = interaction.get("timestamp", "")

                    ws.cell(row=current_row, column=2, value="user:")
                    ws.cell(row=current_row, column=3, value=user_text)
                    current_row += 1

                    ws.cell(row=current_row, column=2, value="assistant:")
                    ws.cell(row=current_row, column=3, value=assistant_text)
                    current_row += 1

                    ws.cell(row=current_row, column=2, value="timestamp:")
                    ws.cell(row=current_row, column=3, value=timestamp_text)
                    current_row += 2

                current_row += 1

            wb.save(output_path)
            logger.info(f"XLSX file created successfully: {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error creating XLSX file: {e}")
            return False

    def _create_pdf(self, data, output_path):
        """
        Create a PDF file from chat history data.

        Args:
            data (list[dict]): Chat history data
            output_path (Path): Path to the output PDF file

        Returns:
            bool: True if successful, False otherwise
        """
        logger.info(f"Creating PDF file: {output_path}")

        try:
            try:
                pdfmetrics.registerFont(
                    TTFont("ArabicFont", FONTS_DIR / "Amiri-Regular.ttf")
                )
                pdfmetrics.registerFont(
                    TTFont("EnglishFont", FONTS_DIR / "Helvetica.ttf")
                )
            except Exception as e:
                logger.error(f"Error registering fonts: {e}")
                return False

            pdf = SimpleDocTemplate(str(output_path), pagesize=landscape(letter))
            styles = getSampleStyleSheet()

            english_style = ParagraphStyle(
                "EnglishStyle",
                parent=styles["Normal"],
                fontName="EnglishFont",
                alignment=TA_LEFT,
                fontSize=12,
                leading=14,
                wordWrap=True,
            )
            arabic_style = ParagraphStyle(
                "ArabicStyle",
                parent=styles["Normal"],
                fontName="ArabicFont",
                alignment=TA_RIGHT,
                fontSize=12,
                leading=14,
                wordWrap=True,
            )

            elements = []
            table_data = self._prepare_pdf_table_data(
                data, english_style, arabic_style, styles
            )

            for i in range(0, len(table_data), 10):
                page_data = table_data[:1] + table_data[1 + i : 1 + i + 10]
                table = Table(
                    page_data,
                    repeatRows=1,
                    hAlign="CENTER",
                    colWidths=[1 * inch, 2 * inch, 4 * inch],
                )
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (0, -1), "LEFT"),
                            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ]
                    )
                )
                elements.append(table)

                if i + 10 < len(table_data):
                    elements.append(PageBreak())

            pdf.build(elements)
            logger.info(f"PDF file created successfully: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error creating PDF file: {e}")
            return False

    def _prepare_pdf_table_data(self, data, english_style, arabic_style, styles):
        """
        Prepare table data for PDF generation.

        Args:
            data (list[dict]): Chat history data
            english_style (ParagraphStyle): Style for English text
            arabic_style (ParagraphStyle): Style for Arabic text
            styles (StyleSheet): ReportLab stylesheet

        Returns:
            list: Table data for PDF generation
        """
        table_data = []

        header = [
            Paragraph("User ID", styles["Heading3"]),
            Paragraph("User", styles["Heading3"]),
            Paragraph("Assistant", styles["Heading3"]),
        ]
        table_data.append(header)

        for user_entry in data:
            user_id = user_entry.get("user_id", "No ID")
            interactions = user_entry.get("interactions", [])

            if interactions:
                first_interaction = interactions[0]
                row = [
                    Paragraph(user_id, english_style),
                    Paragraph(
                        self._prepare_arabic_text(first_interaction.get("user", "")),
                        arabic_style,
                    ),
                    Paragraph(
                        self._prepare_arabic_text(
                            first_interaction.get("assistant", "")
                        ),
                        arabic_style,
                    ),
                ]
                table_data.append(row)

            for interaction in interactions[1:]:
                row = [
                    Paragraph("", english_style),
                    Paragraph(
                        self._prepare_arabic_text(interaction.get("user", "")),
                        arabic_style,
                    ),
                    Paragraph(
                        self._prepare_arabic_text(interaction.get("assistant", "")),
                        arabic_style,
                    ),
                ]
                table_data.append(row)

        return table_data

    @staticmethod
    def _prepare_arabic_text(text):
        """
        Prepare Arabic text for PDF display.

        Args:
            text (str): Text to prepare

        Returns:
            str: Prepared text with proper bidirectional formatting
        """
        if not text:
            return ""

        text = text.replace("\n", " ")

        max_length = 1000
        if len(text) > max_length:
            text = text[:max_length] + "..."

        reshaped_text = arabic_reshaper.reshape(text)
        return get_display(reshaped_text)
